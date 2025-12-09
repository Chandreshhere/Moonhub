#!/usr/bin/env python3
"""
Multi-Platform Inventory Management System
Supports Amazon, Flipkart, Meesho, and other platforms
"""

import json
from datetime import datetime, timedelta
import requests
from typing import Dict, List, Optional
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from database import get_connection, execute_query, setup_database

class InventoryManager:
    def __init__(self, db_path: str = None):
        setup_database()
        self.setup_logging()
        
    def setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    

    
    def add_product(self, sku: str, name: str, category: str, cost_price: float, 
                   selling_price: float, initial_stock: int = 0) -> bool:
        """Add new product to inventory"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO products (sku, name, category, cost_price, selling_price, current_stock)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
            ''' if os.environ.get('DATABASE_URL') else '''
                INSERT INTO products (sku, name, category, cost_price, selling_price, current_stock)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (sku, name, category, cost_price, selling_price, initial_stock))
            
            if os.environ.get('DATABASE_URL'):
                product_id = cursor.fetchone()['id']
            else:
                product_id = cursor.lastrowid
            
            if initial_stock > 0:
                cursor.execute('''
                    INSERT INTO stock_movements (product_id, movement_type, quantity, notes)
                    VALUES (%s, 'IN', %s, 'Initial stock')
                ''' if os.environ.get('DATABASE_URL') else '''
                    INSERT INTO stock_movements (product_id, movement_type, quantity, notes)
                    VALUES (?, 'IN', ?, 'Initial stock')
                ''', (product_id, initial_stock))
            
            conn.commit()
            conn.close()
            self.logger.info(f"Product {sku} added successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding product: {e}")
            return False
    
    def update_stock(self, sku: str, quantity: int, movement_type: str = 'ADJUSTMENT', 
                    platform: str = None, order_id: str = None, notes: str = None) -> bool:
        """Update stock levels"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get product ID and current stock
            param = '%s' if os.environ.get('DATABASE_URL') else '?'
            cursor.execute(f'SELECT id, current_stock FROM products WHERE sku = {param}', (sku,))
            result = cursor.fetchone()
            
            if not result:
                self.logger.error(f"Product {sku} not found")
                conn.close()
                return False
            
            product_id = result['id'] if isinstance(result, dict) else result[0]
            current_stock = result['current_stock'] if isinstance(result, dict) else result[1]
            
            if movement_type == 'OUT' and current_stock < quantity:
                self.logger.error(f"Insufficient stock for {sku}")
                conn.close()
                return False
            
            # Calculate new stock
            if movement_type == 'IN' or movement_type == 'ADJUSTMENT':
                new_stock = current_stock + quantity
            else:
                new_stock = current_stock - quantity
            
            # Update product stock
            cursor.execute(f'''
                UPDATE products SET current_stock = {param}, updated_at = CURRENT_TIMESTAMP 
                WHERE id = {param}
            ''', (new_stock, product_id))
            
            # Record movement
            cursor.execute(f'''
                INSERT INTO stock_movements (product_id, movement_type, quantity, platform, order_id, notes)
                VALUES ({param}, {param}, {param}, {param}, {param}, {param})
            ''', (product_id, movement_type, quantity, platform, order_id, notes))
            
            conn.commit()
            conn.close()
            self.logger.info(f"Stock updated for {sku}: {current_stock} -> {new_stock}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating stock: {e}")
            return False
    
    def add_platform_listing(self, sku: str, platform: str, platform_sku: str, 
                           platform_price: float, stock_allocated: int = 0) -> bool:
        """Add product listing to a platform"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            param = '%s' if os.environ.get('DATABASE_URL') else '?'
            
            cursor.execute(f'SELECT id FROM products WHERE sku = {param}', (sku,))
            result = cursor.fetchone()
            
            if not result:
                self.logger.error(f"Product {sku} not found")
                conn.close()
                return False
            
            product_id = result['id'] if isinstance(result, dict) else result[0]
            
            cursor.execute(f'''
                INSERT INTO platform_listings (product_id, platform, platform_sku, platform_price, stock_allocated)
                VALUES ({param}, {param}, {param}, {param}, {param})
            ''', (product_id, platform, platform_sku, platform_price, stock_allocated))
            
            conn.commit()
            conn.close()
            self.logger.info(f"Platform listing added: {sku} on {platform}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding platform listing: {e}")
            return False
    
    def get_low_stock_alerts(self) -> List[Dict]:
        """Get products with stock below minimum level"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT sku, name, current_stock, min_stock_level
            FROM products
            WHERE current_stock <= min_stock_level
        ''')
        
        results = cursor.fetchall()
        conn.close()
        
        if results and isinstance(results[0], dict):
            return results
        
        return [
            {
                'sku': row[0],
                'name': row[1],
                'current_stock': row[2],
                'min_stock_level': row[3]
            }
            for row in results
        ]
    
    def generate_inventory_report(self) -> List[Dict]:
        """Generate comprehensive inventory report"""
        conn = get_connection()
        cursor = conn.cursor()
        
        if os.environ.get('DATABASE_URL'):
            query = '''
                SELECT 
                    p.sku,
                    p.name,
                    p.category,
                    p.current_stock,
                    p.min_stock_level,
                    p.cost_price,
                    p.selling_price,
                    COUNT(pl.id) as platform_count,
                    STRING_AGG(pl.platform, ',') as platforms
                FROM products p
                LEFT JOIN platform_listings pl ON p.id = pl.product_id
                GROUP BY p.id, p.sku, p.name, p.category, p.current_stock, p.min_stock_level, p.cost_price, p.selling_price
            '''
        else:
            query = '''
                SELECT 
                    p.sku,
                    p.name,
                    p.category,
                    p.current_stock,
                    p.min_stock_level,
                    p.cost_price,
                    p.selling_price,
                    COUNT(pl.id) as platform_count,
                    GROUP_CONCAT(pl.platform) as platforms
                FROM products p
                LEFT JOIN platform_listings pl ON p.id = pl.product_id
                GROUP BY p.id
            '''
        
        cursor.execute(query)
        
        if hasattr(cursor, 'fetchall'):
            results = cursor.fetchall()
            if results and isinstance(results[0], dict):
                conn.close()
                return results
            columns = [desc[0] for desc in cursor.description]
            results = [dict(zip(columns, row)) for row in results]
        
        conn.close()
        return results
    
    def export_to_excel(self, filename: str = None) -> str:
        """Export inventory data to Excel"""
        if not filename:
            filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        tmpdir = os.environ.get('TMPDIR', '/tmp')
        filepath = os.path.join(tmpdir, filename)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Main inventory report
        ws1 = wb.create_sheet('Inventory')
        inventory_data = self.generate_inventory_report()
        if inventory_data:
            headers = list(inventory_data[0].keys())
            ws1.append(headers)
            for row in inventory_data:
                ws1.append([row[h] for h in headers])
        
        # Low stock alerts
        ws2 = wb.create_sheet('Low Stock Alerts')
        low_stock = self.get_low_stock_alerts()
        if low_stock:
            headers = list(low_stock[0].keys())
            ws2.append(headers)
            for row in low_stock:
                ws2.append([row[h] for h in headers])
        
        # Platform listings
        ws3 = wb.create_sheet('Platform Listings')
        conn = get_connection()
        cursor = conn.cursor()
        param = '%s' if os.environ.get('DATABASE_URL') else '?'
        cursor.execute('''
            SELECT 
                p.sku,
                p.name,
                pl.platform,
                pl.platform_sku,
                pl.platform_price,
                pl.stock_allocated,
                pl.is_active
            FROM products p
            JOIN platform_listings pl ON p.id = pl.product_id
        ''')
        
        results = cursor.fetchall()
        if results:
            if isinstance(results[0], dict):
                headers = list(results[0].keys())
                ws3.append(headers)
                for row in results:
                    ws3.append([row[h] for h in headers])
            else:
                columns = [desc[0] for desc in cursor.description]
                ws3.append(columns)
                for row in results:
                    ws3.append(row)
        conn.close()
        
        wb.save(filepath)
        self.logger.info(f"Excel report exported: {filepath}")
        return filepath

# Platform-specific integrations (mock implementations)
class PlatformIntegrator:
    def __init__(self, inventory_manager: InventoryManager):
        self.inventory_manager = inventory_manager
        self.load_platform_configs()
    
    def load_platform_configs(self):
        """Load platform API configurations"""
        try:
            with open('/Users/moon/Documents/inventory/platform_config.json', 'r') as f:
                self.configs = json.load(f)
        except FileNotFoundError:
            self.configs = {}
    
    def sync_amazon_inventory(self):
        """Sync inventory with Amazon (mock implementation)"""
        # This would integrate with Amazon MWS/SP-API
        self.inventory_manager.logger.info("Amazon inventory sync completed")
    
    def sync_flipkart_inventory(self):
        """Sync inventory with Flipkart (mock implementation)"""
        # This would integrate with Flipkart API
        self.inventory_manager.logger.info("Flipkart inventory sync completed")
    
    def sync_meesho_inventory(self):
        """Sync inventory with Meesho (mock implementation)"""
        # This would integrate with Meesho API
        self.inventory_manager.logger.info("Meesho inventory sync completed")

if __name__ == "__main__":
    # Example usage
    inventory = InventoryManager()
    
    # Add sample products
    inventory.add_product("SKU001", "Wireless Headphones", "Electronics", 500, 1200, 50)
    inventory.add_product("SKU002", "Phone Case", "Accessories", 50, 200, 100)
    
    # Add platform listings
    inventory.add_platform_listing("SKU001", "Amazon", "AMZN-WH001", 1199, 20)
    inventory.add_platform_listing("SKU001", "Flipkart", "FK-WH001", 1150, 15)
    
    # Generate report
    report_file = inventory.export_to_excel()
    print(f"Inventory report generated: {report_file}")