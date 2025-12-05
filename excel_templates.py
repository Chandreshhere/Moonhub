#!/usr/bin/env python3
"""
Excel Template Generator for Inventory Management
Creates automated Excel templates with formulas and formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class ExcelTemplateGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_master_inventory_template(self, filename: str = "Master_Inventory_Template.xlsx"):
        """Create comprehensive inventory management template"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Master Inventory Sheet
        ws_inventory = wb.create_sheet("Master Inventory")
        
        headers = [
            "SKU", "Product Name", "Category", "Cost Price", "Selling Price",
            "Current Stock", "Min Stock", "Max Stock", "Reorder Point",
            "Amazon Stock", "Flipkart Stock", "Meesho Stock", "Other Platform Stock",
            "Total Allocated", "Available Stock", "Stock Value", "Profit Margin %",
            "Last Updated", "Status"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws_inventory.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data and formulas
        sample_data = [
            ["SKU001", "Wireless Headphones", "Electronics", 500, 1200, 100, 10, 200, 20, 30, 25, 20, 5, "=J2+K2+L2+M2", "=F2-N2", "=F2*D2", "=(E2-D2)/E2*100", datetime.now().strftime("%Y-%m-%d"), "Active"],
            ["SKU002", "Phone Case", "Accessories", 50, 200, 150, 15, 300, 30, 40, 35, 30, 10, "=J3+K3+L3+M3", "=F3-N3", "=F3*D3", "=(E3-D3)/E3*100", datetime.now().strftime("%Y-%m-%d"), "Active"]
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_inventory.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply conditional formatting
        # Low stock alert (red if current stock <= min stock)
        low_stock_rule = CellIsRule(operator='lessThanOrEqual', formula=['$G2'], fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"))
        ws_inventory.conditional_formatting.add('F2:F1000', low_stock_rule)
        
        # High profit margin (green if > 50%)
        high_profit_rule = CellIsRule(operator='greaterThan', formula=[50], fill=PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"))
        ws_inventory.conditional_formatting.add('Q2:Q1000', high_profit_rule)
        
        # 2. Platform Sync Sheet
        ws_platforms = wb.create_sheet("Platform Sync")
        
        platform_headers = [
            "SKU", "Product Name", "Platform", "Platform SKU", "Platform Price",
            "Stock Allocated", "Last Sync", "Sync Status", "Error Message"
        ]
        
        for col, header in enumerate(platform_headers, 1):
            cell = ws_platforms.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # 3. Stock Movement Log
        ws_movements = wb.create_sheet("Stock Movements")
        
        movement_headers = [
            "Date", "SKU", "Product Name", "Movement Type", "Quantity",
            "Platform", "Order ID", "Previous Stock", "New Stock", "Notes"
        ]
        
        for col, header in enumerate(movement_headers, 1):
            cell = ws_movements.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # 4. Low Stock Alerts
        ws_alerts = wb.create_sheet("Low Stock Alerts")
        
        alert_headers = [
            "SKU", "Product Name", "Current Stock", "Min Stock", "Shortage",
            "Suggested Reorder", "Priority", "Action Required"
        ]
        
        for col, header in enumerate(alert_headers, 1):
            cell = ws_alerts.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add formula for shortage calculation
        ws_alerts.cell(row=2, column=5, value="=IF(C2<D2,D2-C2,0)")
        ws_alerts.cell(row=2, column=6, value="=IF(E2>0,E2*2,0)")
        ws_alerts.cell(row=2, column=7, value='=IF(E2>0,IF(E2>50,"HIGH",IF(E2>20,"MEDIUM","LOW")),"OK")')
        
        # 5. Dashboard
        ws_dashboard = wb.create_sheet("Dashboard")
        
        # Dashboard metrics
        dashboard_data = [
            ["INVENTORY DASHBOARD", ""],
            ["", ""],
            ["Total Products", "=COUNTA('Master Inventory'!A:A)-1"],
            ["Total Stock Value", "=SUM('Master Inventory'!P:P)"],
            ["Low Stock Items", "=COUNTIF('Master Inventory'!F:F,\"<\"&'Master Inventory'!G:G)"],
            ["Out of Stock Items", "=COUNTIF('Master Inventory'!F:F,0)"],
            ["Average Profit Margin", "=AVERAGE('Master Inventory'!Q:Q)"],
            ["", ""],
            ["PLATFORM SUMMARY", ""],
            ["Amazon Listings", "=COUNTIF('Platform Sync'!C:C,\"Amazon\")"],
            ["Flipkart Listings", "=COUNTIF('Platform Sync'!C:C,\"Flipkart\")"],
            ["Meesho Listings", "=COUNTIF('Platform Sync'!C:C,\"Meesho\")"]
        ]
        
        for row_idx, (label, formula) in enumerate(dashboard_data, 1):
            ws_dashboard.cell(row=row_idx, column=1, value=label)
            if formula:
                ws_dashboard.cell(row=row_idx, column=2, value=formula)
        
        # Format dashboard
        ws_dashboard.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws_dashboard.cell(row=9, column=1).font = Font(bold=True, size=14)
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        filepath = f"/Users/moon/Documents/inventory/{filename}"
        wb.save(filepath)
        return filepath
    
    def create_order_processing_template(self, filename: str = "Order_Processing_Template.xlsx"):
        """Create order processing and fulfillment template"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Orders sheet
        ws_orders = wb.create_sheet("Orders")
        
        order_headers = [
            "Order Date", "Order ID", "Platform", "SKU", "Product Name",
            "Quantity", "Unit Price", "Total Amount", "Customer Name",
            "Shipping Address", "Status", "Tracking Number", "Notes"
        ]
        
        for col, header in enumerate(order_headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add data validation for status
        from openpyxl.worksheet.datavalidation import DataValidation
        
        status_validation = DataValidation(
            type="list",
            formula1='"Pending,Processing,Shipped,Delivered,Cancelled,Returned"',
            allow_blank=False
        )
        ws_orders.add_data_validation(status_validation)
        status_validation.add('K2:K1000')
        
        # Fulfillment tracking
        ws_fulfillment = wb.create_sheet("Fulfillment")
        
        fulfillment_headers = [
            "Date", "Order ID", "SKU", "Quantity Picked", "Quantity Packed",
            "Quantity Shipped", "Picker Name", "Packer Name", "Shipping Method",
            "Estimated Delivery", "Status"
        ]
        
        for col, header in enumerate(fulfillment_headers, 1):
            cell = ws_fulfillment.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        filepath = f"/Users/moon/Documents/inventory/{filename}"
        wb.save(filepath)
        return filepath
    
    def create_supplier_management_template(self, filename: str = "Supplier_Management_Template.xlsx"):
        """Create supplier and purchase order template"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Suppliers sheet
        ws_suppliers = wb.create_sheet("Suppliers")
        
        supplier_headers = [
            "Supplier ID", "Supplier Name", "Contact Person", "Email",
            "Phone", "Address", "Payment Terms", "Lead Time (Days)",
            "Rating", "Status", "Notes"
        ]
        
        for col, header in enumerate(supplier_headers, 1):
            cell = ws_suppliers.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Purchase Orders sheet
        ws_po = wb.create_sheet("Purchase Orders")
        
        po_headers = [
            "PO Number", "Date", "Supplier", "SKU", "Product Name",
            "Quantity Ordered", "Unit Cost", "Total Cost", "Expected Delivery",
            "Status", "Received Quantity", "Pending Quantity"
        ]
        
        for col, header in enumerate(po_headers, 1):
            cell = ws_po.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add formulas
        ws_po.cell(row=2, column=8, value="=F2*G2")  # Total Cost
        ws_po.cell(row=2, column=12, value="=F2-K2")  # Pending Quantity
        
        filepath = f"/Users/moon/Documents/inventory/{filename}"
        wb.save(filepath)
        return filepath

if __name__ == "__main__":
    generator = ExcelTemplateGenerator()
    
    # Generate all templates
    templates = [
        generator.create_master_inventory_template(),
        generator.create_order_processing_template(),
        generator.create_supplier_management_template()
    ]
    
    print("Excel templates created:")
    for template in templates:
        print(f"- {template}")